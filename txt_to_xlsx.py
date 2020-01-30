import  openpyxl as px

#ファイルを開く
path = r"G:\atom\igarashi\counter_round10.txt"
file = open(path)

#すべての行をリストとして読み込み
lines = file.readlines()

#ファイルを閉じる
file.close()

#ブックを新規作成。Workbookの'w'は必ず大文字。
wb = px.Workbook()

#出力ファイル名を指定
filename = "result.xlsx"

#アクティブなシートを取得
ws = wb.active

#1行目に2列目から17列目まで1~15の値をセルに代入
for a in range(1, 17):
    ws.cell(row=a+1, column=1).value = a-1

"""
以下、テキストファイルから取得した値を、想定したデザイン通りになるように
セルに書き込んでいく。
"""
#listを取得するための変数
l = 0

#セルへの書き込み
for i in range(2,130):
    for j in range(1,18):
        blank = lines[l].split()
        ws.cell(row=j, column=i).value = int(blank[len(blank)-1])
        l += 1

#ブックを保存する
wb.save(filename)
